import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Bóc tách VT Lan can"

# ========== STYLES ==========
header_font = Font(name='Times New Roman', bold=True, size=12)
title_font = Font(name='Times New Roman', bold=True, size=14)
data_font = Font(name='Times New Roman', size=11)
bold_data_font = Font(name='Times New Roman', bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(name='Times New Roman', bold=True, size=11, color='FFFFFF')
group_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

# ========== COLUMN WIDTHS ==========
col_widths = {1: 6, 2: 45, 3: 10, 4: 18, 5: 12, 6: 18, 7: 14, 8: 16}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ========== TITLE ==========
ws.merge_cells('A1:H1')
title_cell = ws['A1']
title_cell.value = "BẢNG BÓC TÁCH KHỐI LƯỢNG VẬT TƯ - HẠNG MỤC LAN CAN INOX"
title_cell.font = title_font
title_cell.alignment = center_align

ws.merge_cells('A2:H2')
ws['A2'].value = "(Vật liệu: Inox 304)"
ws['A2'].font = Font(name='Times New Roman', italic=True, size=11)
ws['A2'].alignment = center_align

# ========== HEADER ROW ==========
headers = [
    ("STT", 4),
    ("TÊN HẠNG MỤC,\nQUY CÁCH VẬT TƯ", 4),
    ("ĐƠN VỊ\nTÍNH", 4),
    ("KHỐI LƯỢNG\n/ TẦNG", 4),
    ("SỐ\nTẦNG", 4),
    ("TỔNG KHỐI\nLƯỢNG", 4),
    ("TỶ TRỌNG\n(kg/md)", 4),
    ("TỔNG TL\n(kg)", 4),
]

row_start = 4
for col_idx, (header_text, _) in enumerate(headers, 1):
    cell = ws.cell(row=row_start, column=col_idx, value=header_text)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

ws.row_dimensions[row_start].height = 45

# ========== TỶ TRỌNG (kg/md) ==========
# Inox hộp 30x60x1.5mm: ~2.07 kg/m
# Inox hộp 10x40x1.2mm: ~0.91 kg/m
# Tấm inox 40x8mm: ~2.54 kg/m

ty_trong = {
    'hop_30x60': 2.07,
    'hop_10x40': 0.91,
    'tam_40x8': 2.54,
}

# ========== DATA ==========
# Format: (stt, ten_hang_muc, don_vi, kl_tang, so_tang, ty_trong_val, is_group)
# If is_group = True, it's a group header row
# For ban ma: ty_trong_val = weight per piece (kg/cai), don_vi = 'cai'

data = []

# Tỷ trọng tính toán:
# Inox hộp 30x60x1.5: Chu vi = 2*(30+60-2*1.5)=174mm, A=174*1.5=261mm²; W = 261*7.93/1000*10 = 2.07 kg/m
# Inox hộp 10x40x1.2: Chu vi = 2*(10+40-2*1.2)=93.2mm, A=93.2*1.2=111.84mm²; W = 111.84*7.93/1000*10 = 0.89 kg/m  
# Tấm 40x8: A=40*8=320mm²; W = 320*7.93/1000*10 = 2.54 kg/m
# Bản mã 60x70x5: V=60*70*5=21000mm³=21cm³; W = 21*7.93 = 166.53g = 0.167 kg/cái

# ===== LAN CAN 1 =====
data.append(("I", "LAN CAN 1", None, None, None, None, None, True))

# Panel chính (mặt trước) - 4465mm
data.append(("1", "LAN CAN 1 - Panel chính (L=4465mm)", None, None, None, None, None, True))
data.append(("1.1", "Inox hộp 30x60x1.5 (tay vịn, khung)", "md", 4.465, 1, None, 2.07, False))
data.append(("1.2", "Inox hộp 10x40x1.2 (song đứng)", "md", 27.1, 1, None, 0.91, False))
data.append(("1.3", "Tấm inox 40x8 (thanh ngang đế)", "md", 4.445, 1, None, 2.54, False))
data.append(("1.4", "Bản mã inox 60x70x5", "cái", 2, 1, None, 0.167, False))

# Panel hồi (mặt bên) - 1060mm
data.append(("2", "LAN CAN 1 - Panel hồi (L=1060mm)", None, None, None, None, None, True))
data.append(("2.1", "Inox hộp 30x60x1.5 (tay vịn, khung)", "md", 1.06, 1, None, 2.07, False))
data.append(("2.2", "Inox hộp 10x40x1.2 (song đứng)", "md", 6.5, 1, None, 0.91, False))
data.append(("2.3", "Tấm inox 40x8 (thanh ngang đế)", "md", 1.04, 1, None, 2.54, False))

# ===== LAN CAN 2 =====
data.append(("II", "LAN CAN 2", None, None, None, None, None, True))

# Panel chính (mặt trước) - 5645mm
data.append(("3", "LAN CAN 2 - Panel chính (L=5645mm)", None, None, None, None, None, True))
data.append(("3.1", "Inox hộp 30x60x1.5 (tay vịn, khung)", "md", 5.645, 1, None, 2.07, False))
data.append(("3.2", "Inox hộp 10x40x1.2 (song đứng)", "md", 30.0, 1, None, 0.91, False))
data.append(("3.3", "Tấm inox 40x8 (thanh ngang đế)", "md", 5.625, 1, None, 2.54, False))

# Panel hồi 1 (bên trái) - 1654mm
data.append(("4", "LAN CAN 2 - Panel hồi 1 (L=1654mm)", None, None, None, None, None, True))
data.append(("4.1", "Inox hộp 30x60x1.5 (tay vịn, khung)", "md", 1.654, 1, None, 2.07, False))
data.append(("4.2", "Inox hộp 10x40x1.2 (song đứng)", "md", 6.5, 1, None, 0.91, False))
data.append(("4.3", "Tấm inox 40x8 (thanh ngang đế)", "md", 2.0, 1, None, 2.54, False))

# Panel hồi 2 (bên phải) - 1720mm
data.append(("5", "LAN CAN 2 - Panel hồi 2 (L=1720mm)", None, None, None, None, None, True))
data.append(("5.1", "Inox hộp 30x60x1.5 (tay vịn, khung)", "md", 1.72, 1, None, 2.07, False))
data.append(("5.2", "Inox hộp 10x40x1.2 (song đứng)", "md", 9.0, 1, None, 0.91, False))
data.append(("5.3", "Tấm inox 40x8 (thanh ngang đế)", "md", 1.7, 1, None, 2.54, False))
data.append(("5.4", "Bản mã inox 60x70x5", "cái", 3, 1, None, 0.167, False))

# ===== LAN CAN 3 =====
data.append(("III", "LAN CAN 3", None, None, None, None, None, True))

# Panel thẳng - 7310mm (Image 4)
data.append(("6", "LAN CAN 3 - Panel thẳng (L=7310mm)", None, None, None, None, None, True))
data.append(("6.1", "Inox hộp 30x60x1.5 (tay vịn, khung)", "md", 7.31, 1, None, 2.07, False))
data.append(("6.2", "Inox hộp 10x40x1.2 (song đứng)", "md", 29.5, 1, None, 0.91, False))
data.append(("6.3", "Tấm inox 40x8 (thanh ngang đế)", "md", 7.29, 1, None, 2.54, False))

# Panel chữ U - 5115mm (Image 5 main)
data.append(("7", "LAN CAN 3 - Panel chữ U (L=5115mm)", None, None, None, None, None, True))
data.append(("7.1", "Inox hộp 30x60x1.5 (tay vịn, khung)", "md", 6.116, 1, None, 2.07, False))
data.append(("7.2", "Inox hộp 10x40x1.2 (song đứng)", "md", 41.5, 1, None, 0.91, False))
data.append(("7.3", "Tấm inox 40x8 (thanh ngang đế)", "md", 5.095, 1, None, 2.54, False))

# Panel phụ - 846mm (Image 5 side detail)
data.append(("8", "LAN CAN 3 - Panel phụ (L=846mm)", None, None, None, None, None, True))
data.append(("8.1", "Inox hộp 30x60x1.5 (tay vịn, khung)", "md", 0.846, 1, None, 2.07, False))
data.append(("8.2", "Inox hộp 10x40x1.2 (song đứng)", "md", 3.2, 1, None, 0.91, False))
data.append(("8.3", "Tấm inox 40x8 (thanh ngang đế)", "md", 0.826, 1, None, 2.54, False))
data.append(("8.4", "Bản mã inox 60x70x5", "cái", 2, 1, None, 0.167, False))

# ========== WRITE DATA ==========
current_row = row_start + 1

for item in data:
    stt, name, don_vi, kl_tang, so_tang, tong_kl, ty_trong_val, is_group = item
    
    row = current_row
    
    if is_group:
        # Group header row
        ws.cell(row=row, column=1, value=stt).font = bold_data_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=name).font = bold_data_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        
        for col in range(3, 9):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).font = bold_data_font
        
        # Apply group fill
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = group_fill
    else:
        # Data row
        tong_kl_calc = round(kl_tang * so_tang, 3) if kl_tang and so_tang else 0
        tong_tl_calc = round(tong_kl_calc * ty_trong_val, 2) if tong_kl_calc and ty_trong_val else 0
        
        ws.cell(row=row, column=1, value=stt).font = data_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border
        
        ws.cell(row=row, column=2, value=name).font = data_font
        ws.cell(row=row, column=2).alignment = left_align
        ws.cell(row=row, column=2).border = thin_border
        
        ws.cell(row=row, column=3, value=don_vi).font = data_font
        ws.cell(row=row, column=3).alignment = center_align
        ws.cell(row=row, column=3).border = thin_border
        
        ws.cell(row=row, column=4, value=kl_tang).font = data_font
        ws.cell(row=row, column=4).alignment = center_align
        ws.cell(row=row, column=4).border = thin_border
        ws.cell(row=row, column=4).number_format = '0.000'
        
        ws.cell(row=row, column=5, value=so_tang).font = data_font
        ws.cell(row=row, column=5).alignment = center_align
        ws.cell(row=row, column=5).border = thin_border
        
        ws.cell(row=row, column=6, value=tong_kl_calc).font = data_font
        ws.cell(row=row, column=6).alignment = center_align
        ws.cell(row=row, column=6).border = thin_border
        ws.cell(row=row, column=6).number_format = '0.000'
        
        ws.cell(row=row, column=7, value=ty_trong_val).font = data_font
        ws.cell(row=row, column=7).alignment = center_align
        ws.cell(row=row, column=7).border = thin_border
        ws.cell(row=row, column=7).number_format = '0.00'
        
        ws.cell(row=row, column=8, value=tong_tl_calc).font = data_font
        ws.cell(row=row, column=8).alignment = center_align
        ws.cell(row=row, column=8).border = thin_border
        ws.cell(row=row, column=8).number_format = '0.00'
    
    current_row += 1

# ========== SUMMARY ROW ==========
summary_row = current_row + 1
ws.merge_cells(f'A{summary_row}:F{summary_row}')
ws.cell(row=summary_row, column=1, value="TỔNG CỘNG TRỌNG LƯỢNG (kg)").font = Font(name='Times New Roman', bold=True, size=12, color='FF0000')
ws.cell(row=summary_row, column=1).alignment = center_align
ws.cell(row=summary_row, column=1).border = thin_border
for col in range(2, 9):
    ws.cell(row=summary_row, column=col).border = thin_border

# Calculate total weight
total_weight = 0
for item in data:
    stt, name, don_vi, kl_tang, so_tang, tong_kl, ty_trong_val, is_group = item
    if not is_group and kl_tang and so_tang and ty_trong_val:
        total_weight += round(kl_tang * so_tang * ty_trong_val, 2)

ws.cell(row=summary_row, column=8, value=round(total_weight, 2)).font = Font(name='Times New Roman', bold=True, size=12, color='FF0000')
ws.cell(row=summary_row, column=8).alignment = center_align
ws.cell(row=summary_row, column=8).number_format = '0.00'

# ========== NOTES ==========
note_row = summary_row + 2
ws.merge_cells(f'A{note_row}:H{note_row}')
ws.cell(row=note_row, column=1, value="GHI CHÚ:").font = Font(name='Times New Roman', bold=True, size=11)

notes = [
    "- Tỷ trọng tính theo công thức inox 304 (tỷ trọng 7.93 g/cm³)",
    "- Inox hộp 30x60x1.5: A = 2*(30+60-3)*1.5 = 261 mm² → 2.07 kg/m",
    "- Inox hộp 10x40x1.2: A = 2*(10+40-2.4)*1.2 ≈ 114 mm² → 0.91 kg/m",
    "- Tấm inox 40x8: A = 40*8 = 320 mm2 -> 2.54 kg/m",
    "- Ban ma 60x70x5: V = 60*70*5 = 21000 mm3 = 21 cm3 -> 0.167 kg/cai",
    "- So tang mac dinh = 1, vui long dieu chinh theo thuc te cong trinh",
    "- Khoi luong md (met dai) doc tu ban ve GHI CHU cua tung Lan can",
    "- Ban ma: LC1 = 2 cai, LC2 = 3 cai, LC3 = 2 cai",
]

for i, note in enumerate(notes):
    ws.merge_cells(f'A{note_row+1+i}:H{note_row+1+i}')
    ws.cell(row=note_row+1+i, column=1, value=note).font = Font(name='Times New Roman', italic=True, size=10)

# ========== SUMMARY TABLE BY MATERIAL ==========
sum_start = note_row + len(notes) + 3
ws.merge_cells(f'A{sum_start}:H{sum_start}')
ws.cell(row=sum_start, column=1, value="BẢNG TỔNG HỢP VẬT TƯ THEO LOẠI").font = Font(name='Times New Roman', bold=True, size=12)
ws.cell(row=sum_start, column=1).alignment = center_align

sum_headers = ["STT", "QUY CÁCH VẬT TƯ", "ĐƠN VỊ", "TỔNG SL/CHIỀU DÀI", "", "TỶ TRỌNG (kg/đv)", "", "TỔNG TL (kg)"]
for col_idx, h in enumerate(sum_headers, 1):
    cell = ws.cell(row=sum_start+1, column=col_idx, value=h)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell.alignment = center_align
    cell.border = thin_border

# Tổng hợp theo loại vật tư
total_30x60 = 0
total_10x40 = 0
total_tam = 0
total_ban_ma = 0

for item in data:
    stt, name, don_vi, kl_tang, so_tang, tong_kl, ty_trong_val, is_group = item
    if not is_group and kl_tang and so_tang:
        tong = kl_tang * so_tang
        if ty_trong_val == 2.07:
            total_30x60 += tong
        elif ty_trong_val == 0.91:
            total_10x40 += tong
        elif ty_trong_val == 2.54:
            total_tam += tong
        elif ty_trong_val == 0.167:
            total_ban_ma += tong

summary_data = [
    (1, "Inox hộp 30x60x1.5mm (SUS304)", "md", round(total_30x60, 3), 2.07),
    (2, "Inox hộp 10x40x1.2mm (SUS304)", "md", round(total_10x40, 3), 0.91),
    (3, "Tấm inox 40x8mm (SUS304)", "md", round(total_tam, 3), 2.54),
    (4, "Bản mã inox 60x70x5mm (SUS304)", "cái", int(total_ban_ma), 0.167),
]

for i, (stt_s, name_s, dv, tong_cd, ttr) in enumerate(summary_data):
    r = sum_start + 2 + i
    ws.cell(row=r, column=1, value=stt_s).font = data_font
    ws.cell(row=r, column=1).alignment = center_align
    ws.cell(row=r, column=1).border = thin_border
    
    ws.cell(row=r, column=2, value=name_s).font = data_font
    ws.cell(row=r, column=2).alignment = left_align
    ws.cell(row=r, column=2).border = thin_border
    
    ws.cell(row=r, column=3, value=dv).font = data_font
    ws.cell(row=r, column=3).alignment = center_align
    ws.cell(row=r, column=3).border = thin_border
    
    ws.cell(row=r, column=4, value=tong_cd).font = data_font
    ws.cell(row=r, column=4).alignment = center_align
    ws.cell(row=r, column=4).border = thin_border
    ws.cell(row=r, column=4).number_format = '0.000'
    
    ws.cell(row=r, column=5).border = thin_border
    
    ws.cell(row=r, column=6, value=ttr).font = data_font
    ws.cell(row=r, column=6).alignment = center_align
    ws.cell(row=r, column=6).border = thin_border
    ws.cell(row=r, column=6).number_format = '0.00'
    
    ws.cell(row=r, column=7).border = thin_border
    
    tong_tl_s = round(tong_cd * ttr, 2)
    ws.cell(row=r, column=8, value=tong_tl_s).font = bold_data_font
    ws.cell(row=r, column=8).alignment = center_align
    ws.cell(row=r, column=8).border = thin_border
    ws.cell(row=r, column=8).number_format = '0.00'

# Grand total
grand_r = sum_start + 2 + len(summary_data)
ws.merge_cells(f'A{grand_r}:F{grand_r}')
ws.cell(row=grand_r, column=1, value="TỔNG TRỌNG LƯỢNG TOÀN BỘ").font = Font(name='Times New Roman', bold=True, size=12, color='FF0000')
ws.cell(row=grand_r, column=1).alignment = center_align
for col in range(1, 9):
    ws.cell(row=grand_r, column=col).border = thin_border

grand_total = round(total_30x60*2.07 + total_10x40*0.91 + total_tam*2.54 + total_ban_ma*0.167, 2)
ws.cell(row=grand_r, column=8, value=grand_total).font = Font(name='Times New Roman', bold=True, size=14, color='FF0000')
ws.cell(row=grand_r, column=8).alignment = center_align
ws.cell(row=grand_r, column=8).number_format = '0.00'

# ========== PRINT SETUP ==========
ws.page_setup.orientation = 'landscape'
ws.page_setup.paperSize = ws.PAPERSIZE_A4

# ========== SAVE ==========
output_path = r'd:\ViralWindow_Phan_Mem_Nhom_Kinh\Boc_tach_KL_Lan_can.xlsx'
wb.save(output_path)
print(f"[OK] File Excel da duoc tao thanh cong: {output_path}")
print(f"\nTong hop:")
print(f"   - Inox hop 30x60x1.5: {round(total_30x60, 3)} md -> {round(total_30x60*2.07, 2)} kg")
print(f"   - Inox hop 10x40x1.2: {round(total_10x40, 3)} md -> {round(total_10x40*0.91, 2)} kg")
print(f"   - Tam inox 40x8: {round(total_tam, 3)} md -> {round(total_tam*2.54, 2)} kg")
print(f"   - Ban ma 60x70x5: {int(total_ban_ma)} cai -> {round(total_ban_ma*0.167, 2)} kg")
print(f"   => TONG: {grand_total} kg")
